from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from fastapi import Depends
from app.database import get_db
from app.models.models import SocialSellingMetrica, SDRMetrica, CloserMetrica, Venda, Financeiro
from datetime import datetime
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/comercial", tags=["Upload"])


@router.get("/template/{tipo}")
async def download_template(tipo: str):
    """
    Gera e retorna uma planilha modelo Excel para importação em massa
    tipo: 'social-selling', 'sdr', ou 'closer'
    """
    wb = Workbook()
    ws = wb.active

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    if tipo == "social-selling":
        ws.title = "Social Selling"
        headers = ["Data", "Vendedor", "Ativações", "Conversões", "Leads Gerados"]
        example_data = [
            ["2026-02-01", "Jessica Leopoldino", 250, 3, 2],
            ["2026-02-01", "Artur Gabriel", 180, 2, 1],
            ["2026-02-01", "Karina Carla", 120, 1, 1]
        ]
    elif tipo == "sdr":
        ws.title = "SDR"
        headers = ["Data", "SDR", "Funil", "Leads Recebidos", "Reuniões Agendadas", "Reuniões Realizadas"]
        example_data = [
            ["2026-02-01", "Fernando Dutra", "SS", 3, 2, 2],
            ["2026-02-02", "Fernando Dutra", "Quiz", 1, 1, 1]
        ]
    elif tipo == "closer":
        ws.title = "Closer"
        headers = ["Data", "Closer", "Funil", "Calls Agendadas", "Calls Realizadas", "Vendas", "Booking", "Faturamento Bruto", "Faturamento Líquido"]
        example_data = [
            ["2026-02-01", "Fabio Lima", "SS", 2, 2, 1, 1, 8000, 6000],
            ["2026-02-02", "Mona Garcia", "SS", 1, 1, 0, 0, 0, 0]
        ]
    elif tipo == "vendas":
        ws.title = "Vendas"
        headers = ["Data", "Cliente", "Closer", "Funil", "Tipo Receita", "Produto", "Previsto", "Valor Bruto", "Valor Líquido"]
        example_data = [
            ["2026-02-01", "Dr. João Silva", "Fabio Lima", "Social Selling", "Recorrência", "Assessoria Start", 5000, 8000, 6000],
            ["2026-02-05", "Dra. Maria Santos", "Mona Garcia", "Quiz", "Venda", "Programa Ativ", 12000, 15000, 12000]
        ]
    elif tipo == "financeiro":
        ws.title = "Financeiro"
        headers = ["Data", "Tipo", "Categoria", "Descrição", "Valor", "Previsto/Realizado", "Tipo Custo", "Centro Custo"]
        example_data = [
            ["2026-02-01", "entrada", "Assessoria", "Venda Cliente X", 5000, "realizado", "", ""],
            ["2026-02-05", "saida", "Equipe", "Salário Closer", 3000, "realizado", "Fixo", "Comercial"],
            ["2026-02-10", "saida", "Ferramenta", "Assinatura CRM", 500, "realizado", "Fixo", "Operação"]
        ]
    else:
        raise HTTPException(status_code=400, detail="Tipo inválido. Use: social-selling, sdr, closer, vendas ou financeiro")

    # Adicionar cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Adicionar dados de exemplo
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Salvar em buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=template_{tipo}.xlsx"}
    )


@router.post("/upload/{tipo}")
async def upload_metrics(
    tipo: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Processa upload de planilha Excel com métricas em massa
    tipo: 'social-selling', 'sdr', ou 'closer'
    """
    try:
        # Ler arquivo Excel
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))

        importados = 0
        erros = []

        if tipo == "social-selling":
            # Validar colunas
            required_cols = ["Data", "Vendedor", "Ativações", "Conversões", "Leads Gerados"]
            if not all(col in df.columns for col in required_cols):
                raise HTTPException(status_code=400, detail=f"Planilha deve ter as colunas: {', '.join(required_cols)}")

            for idx, row in df.iterrows():
                try:
                    data = pd.to_datetime(row["Data"]).date()
                    metrica = SocialSellingMetrica(
                        data=data,
                        mes=data.month,
                        ano=data.year,
                        vendedor=str(row["Vendedor"]),
                        ativacoes=int(row["Ativações"]),
                        conversoes=int(row["Conversões"]),
                        leads_gerados=int(row["Leads Gerados"])
                    )
                    db.add(metrica)
                    importados += 1
                except Exception as e:
                    erros.append(f"Linha {idx + 2}: {str(e)}")

        elif tipo == "sdr":
            required_cols = ["Data", "SDR", "Funil", "Leads Recebidos", "Reuniões Agendadas", "Reuniões Realizadas"]
            if not all(col in df.columns for col in required_cols):
                raise HTTPException(status_code=400, detail=f"Planilha deve ter as colunas: {', '.join(required_cols)}")

            for idx, row in df.iterrows():
                try:
                    data = pd.to_datetime(row["Data"]).date()
                    metrica = SDRMetrica(
                        data=data,
                        mes=data.month,
                        ano=data.year,
                        sdr=str(row["SDR"]),
                        funil=str(row["Funil"]),
                        leads_recebidos=int(row["Leads Recebidos"]),
                        reunioes_agendadas=int(row["Reuniões Agendadas"]),
                        reunioes_realizadas=int(row["Reuniões Realizadas"])
                    )
                    db.add(metrica)
                    importados += 1
                except Exception as e:
                    erros.append(f"Linha {idx + 2}: {str(e)}")

        elif tipo == "closer":
            required_cols = ["Data", "Closer", "Funil", "Calls Agendadas", "Calls Realizadas", "Vendas", "Booking", "Faturamento Bruto", "Faturamento Líquido"]
            if not all(col in df.columns for col in required_cols):
                raise HTTPException(status_code=400, detail=f"Planilha deve ter as colunas: {', '.join(required_cols)}")

            for idx, row in df.iterrows():
                try:
                    data = pd.to_datetime(row["Data"]).date()

                    # Ler valores
                    calls_agendadas = int(row["Calls Agendadas"]) if pd.notna(row["Calls Agendadas"]) else 0
                    calls_realizadas = int(row["Calls Realizadas"]) if pd.notna(row["Calls Realizadas"]) else 0
                    vendas = int(row["Vendas"]) if pd.notna(row["Vendas"]) else 0
                    booking = int(row["Booking"]) if pd.notna(row["Booking"]) else 0
                    faturamento_bruto = float(row["Faturamento Bruto"]) if pd.notna(row["Faturamento Bruto"]) else 0.0
                    faturamento_liquido = float(row["Faturamento Líquido"]) if pd.notna(row["Faturamento Líquido"]) else 0.0

                    # Calcular taxas
                    tx_comparecimento = (calls_realizadas / calls_agendadas * 100) if calls_agendadas > 0 else 0.0
                    tx_conversao = (vendas / calls_realizadas * 100) if calls_realizadas > 0 else 0.0
                    ticket_medio = (faturamento_liquido / vendas) if vendas > 0 else 0.0

                    metrica = CloserMetrica(
                        data=data,
                        mes=data.month,
                        ano=data.year,
                        closer=str(row["Closer"]),
                        funil=str(row["Funil"]),
                        calls_agendadas=calls_agendadas,
                        calls_realizadas=calls_realizadas,
                        vendas=vendas,
                        booking=booking,
                        faturamento_bruto=faturamento_bruto,
                        faturamento_liquido=faturamento_liquido,
                        tx_comparecimento=tx_comparecimento,
                        tx_conversao=tx_conversao,
                        ticket_medio=ticket_medio
                    )
                    db.add(metrica)
                    importados += 1
                except Exception as e:
                    erros.append(f"Linha {idx + 2}: {str(e)}")

        elif tipo == "vendas":
            required_cols = ["Data", "Cliente", "Closer", "Funil", "Tipo Receita", "Produto", "Previsto", "Valor Bruto", "Valor Líquido"]
            if not all(col in df.columns for col in required_cols):
                raise HTTPException(status_code=400, detail=f"Planilha deve ter as colunas: {', '.join(required_cols)}")

            for idx, row in df.iterrows():
                try:
                    data = pd.to_datetime(row["Data"]).date()

                    venda = Venda(
                        data=data,
                        mes=data.month,
                        ano=data.year,
                        cliente=str(row["Cliente"]) if pd.notna(row["Cliente"]) else None,
                        closer=str(row["Closer"]) if pd.notna(row["Closer"]) else None,
                        funil=str(row["Funil"]) if pd.notna(row["Funil"]) else None,
                        tipo_receita=str(row["Tipo Receita"]) if pd.notna(row["Tipo Receita"]) else None,
                        produto=str(row["Produto"]) if pd.notna(row["Produto"]) else None,
                        previsto=float(row["Previsto"]) if pd.notna(row["Previsto"]) else 0.0,
                        valor_bruto=float(row["Valor Bruto"]) if pd.notna(row["Valor Bruto"]) else 0.0,
                        valor_liquido=float(row["Valor Líquido"]) if pd.notna(row["Valor Líquido"]) else 0.0,
                    )
                    db.add(venda)
                    importados += 1
                except Exception as e:
                    erros.append(f"Linha {idx + 2}: {str(e)}")

        elif tipo == "financeiro":
            required_cols = ["Data", "Tipo", "Categoria", "Descrição", "Valor", "Previsto/Realizado"]
            if not all(col in df.columns for col in required_cols):
                raise HTTPException(status_code=400, detail=f"Planilha deve ter as colunas: {', '.join(required_cols)}")

            for idx, row in df.iterrows():
                try:
                    data = pd.to_datetime(row["Data"]).date()
                    tipo = str(row["Tipo"]).lower()

                    if tipo not in ['entrada', 'saida']:
                        erros.append(f"Linha {idx + 2}: Tipo deve ser 'entrada' ou 'saida'")
                        continue

                    financeiro = Financeiro(
                        data=data,
                        mes=data.month,
                        ano=data.year,
                        tipo=tipo,
                        categoria=str(row["Categoria"]) if pd.notna(row["Categoria"]) else None,
                        descricao=str(row["Descrição"]) if pd.notna(row["Descrição"]) else None,
                        valor=float(row["Valor"]),
                        previsto_realizado=str(row["Previsto/Realizado"]).lower() if pd.notna(row["Previsto/Realizado"]) else "realizado",
                        tipo_custo=str(row["Tipo Custo"]) if "Tipo Custo" in df.columns and pd.notna(row.get("Tipo Custo")) else None,
                        centro_custo=str(row["Centro Custo"]) if "Centro Custo" in df.columns and pd.notna(row.get("Centro Custo")) else None,
                    )
                    db.add(financeiro)
                    importados += 1
                except Exception as e:
                    erros.append(f"Linha {idx + 2}: {str(e)}")

        else:
            raise HTTPException(status_code=400, detail="Tipo inválido")

        db.commit()

        return {
            "message": "Upload processado com sucesso",
            "importados": importados,
            "erros": len(erros),
            "detalhes_erros": erros  # Retornar todos os erros para debug
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao processar upload: {str(e)}")
